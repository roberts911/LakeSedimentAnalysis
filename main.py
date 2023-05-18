import matplotlib.pyplot as plt
import openpyxl
import tkinter as tk
from tkinter import ttk
import folium
import webbrowser
from geopy.geocoders import Nominatim
from fake_useragent import UserAgent

# Load the Excel file
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active

# Create a list to store lake names
lake_names = []

# Get the lake names and store them in the list lake_names
for row in sheet.iter_rows(min_row=6, min_col=4, values_only=True):
    lake_name = row[0]
    if lake_name is not None:
        lake_names.append(lake_name)

# Create the user interface
root = tk.Tk()
root.title("Elements content in the lake")

# Set the window size
root.geometry("600x400")

# Create the main frame for the content
content_frame = ttk.Frame(root)
content_frame.place(relx=0.5, rely=0.4, anchor=tk.CENTER)

# Create a dropdown for selecting a lake
selected_lake = tk.StringVar()
lake_label = ttk.Label(content_frame, text="Select a lake:")
lake_label.pack(pady=10)
lake_dropdown = ttk.Combobox(content_frame, textvariable=selected_lake, values=lake_names)
lake_dropdown.pack()

# Add a label for the elements selection
elements_label = ttk.Label(content_frame, text="Select elements:")
elements_label.pack(pady=10)

# Create a frame for the CheckBox buttons
checkbox_frame = ttk.Frame(content_frame)
checkbox_frame.pack(pady=10)

# Create CheckBox buttons for each element
element_vars = []
for index, row in enumerate(sheet.iter_rows(min_row=5, max_col=13, max_row=5, values_only=True)):
    for col, element in enumerate(row[7:13]):
        element_var = tk.BooleanVar()
        element_checkbox = ttk.Checkbutton(checkbox_frame, text=element, variable=element_var)
        element_checkbox.grid(row=index, column=col, padx=10)
        element_vars.append(element_var)

def display_chart():
    """
    Display a pie chart showing the content of selected elements in the chosen lake.

    This function retrieves the data for the chosen lake from the Excel file, selects only the 
    chosen elements, and then creates a pie chart showing their content in the lake. 
    The chart also includes a legend and a label with the lake's pH value.
    """
    # Get the name of the selected lake
    lake_name = selected_lake.get()
    # Get the element values for the selected lake
    root_values = []
    root_labels = []
    for row in sheet.iter_rows(min_row=6, max_col=13, values_only=True):
        name = row[3]
        if name == lake_name:
            for cell in row[6:13]:
                root_values.append(cell)

    for row in sheet.iter_rows(min_row=5, max_col=13, max_row=5, values_only=True):
        for cell in row[6:13]:
            root_labels.append(cell)
 
    # Select only the chosen elements
    selected_values = []
    selected_labels = []

    for i, var in enumerate(element_vars):
        if var.get():
            selected_values.append(root_values[i + 1])
            selected_labels.append(root_labels[i + 1])

    # Create the pie chart
    fig, ax = plt.subplots(figsize=(18, 8))
    ax.axis('equal')
    ax.set_title(f'Element content in {lake_name} in 2021')
    selected_explode = [0.4 if label == 'Pb' else 0 for label in selected_labels]
    wedges, labels = ax.pie(selected_values, labels=selected_values, explode=selected_explode)

    # Create a legend
    ax.legend(wedges, selected_labels, title="Elements in mg/kg dm", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))

    # Label with pH under the chart
    ax.text(0, -1.5, f'pH: {root_values[0]}', fontsize=12, ha='center')

    # Display the chart
    plt.show()

# Create a button to display the chart for the selected lake
display_button = ttk.Button(content_frame, text="Display chart", command=display_chart)
display_button.pack(pady=10)

def display_map():
    """
    Display a map centered on the chosen lake, with a marker showing its name and coordinates.

    This function uses the Nominatim geocoder to get the geographic coordinates for the chosen 
    lake, and then creates a map centered on those coordinates using Folium. It adds a marker 
    at the lake's location, with a popup showing the lake's name and coordinates, and also adds 
    an interactive popup that shows the coordinates of the point the user clicks on.
    """
    # Get the name of the selected lake
    lake_name = selected_lake.get()

    # Remove hyphens and everything after them from the lake name
    lake_name = lake_name.split("-")[0].strip()

    # Use geocoder to get the geographic coordinates for the selected lake
    user_agent = UserAgent()
    geolocator = Nominatim(user_agent=user_agent.random)
    location = geolocator.geocode(lake_name)
    if location is not None:
        lat, lon = location.latitude, location.longitude
        # Create a map with the selected lake
        lake_map = folium.Map(location=[lat, lon], zoom_start=13)

        # Add a marker with the lake name and coordinates
        folium.Marker(
            location=[lat, lon],
            popup=lake_name,
            icon=folium.Icon(color='blue')
        ).add_to(lake_map)

        # Add an interactive popup with coordinates
        folium.LatLngPopup().add_to(lake_map)

        # Display the map in the browser
        lake_map.save('lake_map.html')
        webbrowser.open('lake_map.html')
    else:
        print("Failed to find coordinates for this lake.")

# Create a button to display the map for the selected lake
display_map_button = ttk.Button(content_frame, text="Display map", command=display_map)
display_map_button.pack(pady=10)    

def on_close():
    """
    Handle the window close event by printing a message and destroying the root Tkinter object.
    """
    print("Closing the program...")
    root.destroy()
    
# Handle the window close event
root.protocol("WM_DELETE_WINDOW", on_close)

# Run the GUI 
root.mainloop()